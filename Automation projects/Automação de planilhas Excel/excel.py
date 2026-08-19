from openpyxl import load_workbook

#Abre a Planilha

planilha = load_workbook("")#Nome do Arquivo

#Seleciona a Aba Ativa

aba = planilha.active

#Cabeçalho

aba [""] = ""

#Percorre as Linhas

for linha in range(2,aba.max_row +1):
    quantidade = aba[f"B{linha}"].value
    preco = aba[f"C{linha}"].value

    total = quantidade * preco

    aba[f"D{linha}"] = total

#Salva as Alterações

planilha.save("")#Nome do arquivo

print("Planilha atualizada!")